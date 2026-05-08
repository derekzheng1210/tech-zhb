#!/usr/bin/env python3
"""
信用债择券辅助工具 - Flask Web 应用
支持两种数据更新方式:
  1. Git push 新 Excel → Render 自动部署
  2. 访问 /upload 后台上传 Excel → 即时生效
"""
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, jsonify, session
from openpyxl import load_workbook

# ─────────────────────────── 配置 ───────────────────────────
SHEET_NAME = "万得"
DATA_DIR = Path(__file__).parent
DEFAULT_EXCEL = DATA_DIR / "数据.xlsx"
UPLOAD_PASSWORD = "Abcd123%"  # 上传密码，部署前修改

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-in-production")

# ─────────────────────────── 全局数据缓存 ───────────────────────────
BONDS_CACHE = []
DATA_TIMESTAMP = "尚未加载"


# ─────────────────────────── Excel 读取 ───────────────────────────
def read_excel(path: Path) -> list:
    """读取 Excel 并返回紧凑数组"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    bonds = []
    skipped = 0
    for r in rows:
        if not r or len(r) < 11:
            skipped += 1
            continue
        code = str(r[0] or "").strip()
        name = str(r[1] or "").strip()
        term = r[2]
        rating = str(r[3] or "").strip()
        issuer = str(r[4] or "").strip()
        ytm = r[5]
        entity = str(r[6] or "").strip()
        ct = str(r[7] or "").strip()
        sub = str(r[8] or "").strip()
        tech = str(r[9] or "").strip()
        ir = str(r[10] or "").strip()

        if not code or not name or term is None or ytm is None:
            skipped += 1
            continue

        try:
            term = round(float(term), 4)
            ytm = round(float(ytm), 4)
        except (ValueError, TypeError):
            skipped += 1
            continue

        bonds.append([code, name, term, rating, issuer, ytm, entity, ct, sub, tech, ir])

    print(f"读取 {len(bonds)} 条债券数据，跳过 {skipped} 条无效行")
    return bonds


def load_data(excel_path=None):
    """加载/重载数据到全局缓存"""
    global BONDS_CACHE, DATA_TIMESTAMP
    path = Path(excel_path) if excel_path else DEFAULT_EXCEL
    if not path.exists():
        print(f"警告: Excel 文件不存在 {path}")
        BONDS_CACHE = []
        DATA_TIMESTAMP = "数据文件未找到"
        return
    BONDS_CACHE = read_excel(path)
    DATA_TIMESTAMP = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"数据加载完成: {len(BONDS_CACHE)} 条，时间: {DATA_TIMESTAMP}")


# ─────────────────────────── 路由 ───────────────────────────

@app.route("/")
def index():
    """主页面"""
    return render_template("index.html",
                           bond_data=json.dumps(BONDS_CACHE, ensure_ascii=False),
                           timestamp=DATA_TIMESTAMP,
                           total=len(BONDS_CACHE))

@app.route("/upload", methods=["GET", "POST"])
def upload():
    """后台上传管理页面"""
    error = None
    success = None

    if request.method == "POST":
        password = request.form.get("password", "")
        if password != UPLOAD_PASSWORD:
            error = "密码错误"
        elif "excel" not in request.files:
            error = "未选择文件"
        else:
            file = request.files["excel"]
            if file.filename == "":
                error = "未选择文件"
            elif not file.filename.lower().endswith(('.xlsx', '.xls')):
                error = "仅支持 .xlsx / .xls 文件"
            else:
                try:
                    # 备份旧文件
                    if DEFAULT_EXCEL.exists():
                        backup = DATA_DIR / f"数据_备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        DEFAULT_EXCEL.rename(backup)

                    # 保存新文件
                    file.save(str(DEFAULT_EXCEL))

                    # 重载数据
                    load_data()

                    success = f"上传成功！已加载 {len(BONDS_CACHE)} 条债券数据"
                except Exception as e:
                    error = f"处理失败: {str(e)}"

    return render_template("upload.html",
                           error=error,
                           success=success,
                           timestamp=DATA_TIMESTAMP,
                           total=len(BONDS_CACHE))


@app.route("/api/info")
def api_info():
    """返回数据状态信息"""
    return jsonify({
        "total": len(BONDS_CACHE),
        "timestamp": DATA_TIMESTAMP,
    })


# ─────────────────────────── 启动 ───────────────────────────
if __name__ == "__main__":
    # 首次加载数据
    load_data()
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)